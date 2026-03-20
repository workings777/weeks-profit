"""
RPA 통합 자동화 스크립트 - Fursys ERP
기능: 로그인 → ERP Korea → 품목별 수주/판매실적 조회 → 엑셀 다운로드 → 엑셀 가공
사용법: python rpa_fursys_v2.py
"""

import os, glob, traceback
from datetime import datetime
from pathlib import Path
from playwright.sync_api import sync_playwright

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: pip install openpyxl xlrd")
    input("exit..."); exit(1)

# ============================================================
# ✅ 설정 영역
# ============================================================
USERNAME     = "csk830"
PASSWORD     = "sidiz12!!"
DOWNLOAD_DIR = r"C:\Users\FURSYS\Desktop\kyun\weeks_profit\input"
OUTPUT_DIR   = r"C:\Users\FURSYS\Desktop\kyun\weeks_profit\output"
# ============================================================

LOGIN_URL      = "https://ep.fursys.com/account/login.do"
SEARCH_KEYWORD = "품목별 수주/판매실적"
WF = "#mainframe_VFrameSet_HFrameSet_VFrameSet1_workFrame_04006015_form_div_work"


# ============================================================
# 📊 엑셀 가공 함수
# ============================================================
def log(msg, level="INFO"):
    print({"INFO":"V","ERROR":"X","STEP":">","WARN":"!"}.get(level," ") + " " + msg)

def get_latest_excel(folder):
    files = []
    for p in ["*.xlsx","*.xls","*.xlsm"]:
        files.extend(glob.glob(os.path.join(folder, p)))
    return max(files, key=os.path.getmtime) if files else None

def to_float(val):
    try: return float(val) if val is not None else 0.0
    except: return 0.0

def convert_xls(xls_path):
    try:
        import xlrd
    except ImportError:
        print("ERROR: pip install xlrd"); input("exit..."); exit(1)
    log("xls converting", "STEP")
    book = xlrd.open_workbook(xls_path)
    wb = Workbook()
    wb.remove(wb.active)
    for si in range(book.nsheets):
        xs = book.sheet_by_index(si)
        ws = wb.create_sheet(title=xs.name)
        for r in range(xs.nrows):
            for c in range(xs.ncols):
                cell = xs.cell(r, c)
                if cell.ctype == 2:
                    val = int(cell.value) if cell.value == int(cell.value) else cell.value
                elif cell.ctype == 3:
                    val = xlrd.xldate_as_datetime(cell.value, book.datemode)
                elif cell.ctype == 4:
                    val = bool(cell.value)
                elif cell.ctype == 5:
                    val = None
                else:
                    val = cell.value
                ws.cell(row=r+1, column=c+1, value=val)
    log("convert done")
    return wb

def create_pivot(wb, dsn):
    EXCLUDE_SABSO = {"Total", "관계사", "수출사업소", "쌤플", "하자조치내수", "북미사업소"}

    wd = wb[dsn]
    if "pivot" in wb.sheetnames: del wb["pivot"]
    wp = wb.create_sheet("pivot")
    pivot = {}
    for row in range(2, wd.max_row+1):
        b = wd.cell(row=row, column=2).value
        if b is None: continue
        b = str(b).strip()
        if b in EXCLUDE_SABSO: continue
        t = to_float(wd.cell(row=row, column=20).value)
        v = to_float(wd.cell(row=row, column=22).value)
        if b not in pivot: pivot[b] = {"T":0.0,"V":0.0}
        pivot[b]["T"] += t
        pivot[b]["V"] += v
    bd = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
    hf = PatternFill("solid", start_color="4472C4")
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    for col, h in enumerate(["sabso","T_sum","V_sum","V/T"],1):
        c = wp.cell(row=1, column=col, value=h)
        c.fill=hf; c.font=hfont; c.alignment=Alignment(horizontal="center"); c.border=bd
    ef = PatternFill("solid", start_color="DCE6F1")
    for i,(sabso,vals) in enumerate(sorted(pivot.items()),start=2):
        t,v = vals["T"],vals["V"]
        ratio = (v/t) if t!=0 else 0
        fill = ef if i%2==0 else None
        for col,val in enumerate([sabso,t,v,ratio],1):
            c = wp.cell(row=i, column=col, value=val)
            c.border=bd; c.alignment=Alignment(horizontal="center" if col>1 else "left")
            if fill: c.fill=fill
            if col in (2,3): c.number_format='#,##0.00'
            elif col==4: c.number_format='0.00%'
    tr = len(pivot)+2
    tt = sum(v["T"] for v in pivot.values())
    vt = sum(v["V"] for v in pivot.values())
    rt = (vt/tt) if tt!=0 else 0
    tf = PatternFill("solid", start_color="FFC000")
    tfont = Font(bold=True, name="Arial", size=11)
    for col,val in enumerate(["total",tt,vt,rt],1):
        c = wp.cell(row=tr, column=col, value=val)
        c.fill=tf; c.font=tfont; c.border=bd
        c.alignment=Alignment(horizontal="center" if col>1 else "left")
        if col in (2,3): c.number_format='#,##0.00'
        elif col==4: c.number_format='0.00%'
    for col,w in enumerate([20,25,25,20],1):
        wp.column_dimensions[get_column_letter(col)].width=w
    log("pivot done " + str(len(pivot)) + " rows")

def process_excel(src, out):
    log("load: " + os.path.basename(src), "STEP")
    wb = convert_xls(src) if Path(src).suffix.lower()==".xls" else load_workbook(src)
    ws = wb.active
    dsn = ws.title
    mr = ws.max_row
    log("sheet: " + dsn + " rows: " + str(mr))
    for row in ws.iter_rows():
        for cell in row:
            ea = cell.alignment
            cell.alignment = Alignment(wrap_text=True, horizontal=ea.horizontal if ea else None, vertical=ea.vertical if ea else None)
    log("wrap done")
    ws.insert_cols(22)
    hc = ws.cell(row=1, column=22, value="ingoga x qty")
    hc.font=Font(bold=True,name="Arial",size=11)
    hc.fill=PatternFill("solid",start_color="FFF2CC")
    hc.alignment=Alignment(horizontal="center",wrap_text=True)
    cnt=0
    for row in range(2, mr+1):
        if ws.cell(row=row,column=2).value is None: continue
        r_val = to_float(ws.cell(row=row,column=18).value)
        u_val = to_float(ws.cell(row=row,column=21).value)
        c = ws.cell(row=row,column=22,value=r_val*u_val)
        c.number_format='#,##0.00'
        cnt+=1
    log("V done " + str(cnt) + " rows")
    create_pivot(wb, dsn)
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    log("saved: " + out)
    return out

def run_excel():
    print("\n" + "="*50)
    print("📊 PHASE 2: 엑셀 가공 시작")
    print("="*50)
    if not os.path.isdir(DOWNLOAD_DIR):
        log("INPUT 폴더 없음", "ERROR"); return False
    latest = get_latest_excel(DOWNLOAD_DIR)
    if not latest:
        log("엑셀 파일 없음", "ERROR"); return False
    log("파일: " + os.path.basename(latest))
    ft = datetime.fromtimestamp(os.path.getmtime(latest)).strftime("%Y%m%d_%H%M%S")
    out = os.path.join(OUTPUT_DIR, Path(latest).stem + "_done_" + ft + ".xlsx")
    try:
        result = process_excel(latest, out)
        print("\n" + "="*50)
        log("엑셀 가공 완료!")
        log("저장위치: " + result)
        print("="*50)
        return True
    except Exception as e:
        log("엑셀 가공 오류: " + str(e), "ERROR")
        traceback.print_exc()
        return False


# ============================================================
# 🌐 ERP 자동화 함수
# ============================================================
def input_date(prompt, default):
    val = input(f"{prompt} (기본값: {default}, Enter 시 기본값 사용): ").strip()
    return val if val else default

def get_credentials():
    print("\n🔐 ERP 로그인 정보를 입력해 주세요")
    username = input("  아이디: ").strip()
    password = input("  비밀번호: ").strip()
    return username, password

def get_dates():
    today = datetime.today()
    default_from = today.replace(day=1).strftime("%Y-%m-%d")
    default_to   = today.strftime("%Y-%m-%d")
    print("\n📅 조회 기간을 입력해 주세요 (형식: YYYY-MM-DD)")
    date_from = input_date("  시작일", default_from)
    date_to   = input_date("  종료일", default_to)
    print(f"\n  ✅ 조회 기간: {date_from} ~ {date_to}\n")
    return date_from, date_to

def run_erp():
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    today = datetime.today()
    DATE_FROM = today.replace(day=1).strftime("%Y-%m-%d")
    DATE_TO = today.strftime("%Y-%m-%d")
    print(f"\n📅 조회 기간: {DATE_FROM} ~ {DATE_TO}")

    with sync_playwright() as p:
        print("🚀 크롬 브라우저 실행 중...")
        browser = p.chromium.launch(
            channel="chrome",
            headless=False,
            slow_mo=300,
            args=["--start-maximized"]
        )
        context = browser.new_context(viewport=None, accept_downloads=True)
        page = context.new_page()

        try:
            # ── STEP 1: 로그인 ────────────────────────────────
            print(f"\n[1/5] 🌐 로그인")
            page.goto(LOGIN_URL, wait_until="domcontentloaded", timeout=60000)
            page.locator("input[name='userId'], input[type='text']").first.fill(USERNAME)
            page.locator("input[name='userPw'], input[type='password']").first.fill(PASSWORD)
            page.locator("button.btn_full._purple").click()
            page.wait_for_load_state("networkidle")
            if "login" in page.url:
                print("      ⚠️  로그인 실패 - 아이디/비밀번호를 확인해 주세요.")
                return False
            print("      ✅ 로그인 완료")

            # ── STEP 2: 메뉴 버튼 → ERP Korea 클릭 ──────────
            print(f"\n[2/5] 🖱️  ERP Korea 진입")
            menu_btn = page.locator("button.menu[onclick*='main_menu']")
            menu_btn.wait_for(state="visible", timeout=15000)
            menu_btn.click()
            page.wait_for_timeout(1000)

            with context.expect_page(timeout=15000) as new_page_info:
                erp_menu = page.locator("li.menu._ERP:has-text('ERP Korea')")
                erp_menu.wait_for(state="visible", timeout=30000)
                erp_menu.scroll_into_view_if_needed()
                page.wait_for_timeout(500)
                erp_menu.click(force=True)

            erp_page = new_page_info.value
            erp_page.wait_for_load_state("load", timeout=60000)
            erp_page.wait_for_timeout(5000)
            erp_page.evaluate("window.moveTo(0,0); window.resizeTo(screen.width, screen.height);")
            print(f"      ✅ ERP Korea 로딩 완료")

            # ── STEP 3: 화면검색 → 품목별 수주/판매실적 ──────
            print(f"\n[3/5] 🔍 화면검색: '{SEARCH_KEYWORD}'")
            erp_page.get_by_text("화면검색(Menu search)").click()
            erp_page.wait_for_timeout(500)
            erp_page.locator("#mainframe_VFrameSet_topFrame_form_edt_searchMenu_input").fill(SEARCH_KEYWORD)
            erp_page.wait_for_timeout(500)
            erp_page.locator("#mainframe_VFrameSet_topFrame_form_btn_searchMenu > div").click()
            erp_page.wait_for_timeout(1000)
            erp_page.get_by_text(SEARCH_KEYWORD).click()
            erp_page.wait_for_timeout(3000)
            print(f"      ✅ '{SEARCH_KEYWORD}' 화면 진입 완료")

            # ── STEP 4: 조회 조건 설정 ────────────────────────
            print(f"\n[4/5] ⚙️  조회 조건 설정")
            print(f"      📅 기간: {DATE_FROM} ~ {DATE_TO}")

            # 회사 선택
            erp_page.locator(f"{WF}_div_search_cbo_cmpCd_dropbutton > div").click()
            erp_page.wait_for_timeout(300)
            cmp_input = erp_page.locator(f"{WF}_div_search_cbo_cmpCd_comboedit_input")
            cmp_input.press("ArrowDown")
            cmp_input.press("ArrowDown")
            cmp_input.press("ArrowDown")
            cmp_input.press("Enter")
            erp_page.wait_for_timeout(300)
            print(f"      ✅ 회사 선택")

            # 브랜드 선택
            erp_page.locator(f"{WF}_cbo_brdCd_dropbutton > div").click()
            erp_page.wait_for_timeout(300)
            brd_input = erp_page.locator(f"{WF}_cbo_brdCd_comboedit_input")
            brd_input.press("ArrowUp")
            brd_input.press("ArrowDown")
            brd_input.press("Enter")
            erp_page.wait_for_timeout(300)
            print(f"      ✅ 브랜드 선택")

            # 시리즈 전체 선택
            erp_page.locator(f"{WF}_div_search_cbo_ComSeries_dropbutton > div").click()
            erp_page.wait_for_timeout(300)
            series_input = erp_page.locator(f"{WF}_div_search_cbo_ComSeries_comboedit_input")
            series_input.press("Home")
            series_input.press("ArrowDown")
            series_input.press("Enter")
            erp_page.wait_for_timeout(300)
            print(f"      ✅ 시리즈 전체 선택")

            # 날짜 입력 프레임 탐색
            CAL_FROM_ID = "mainframe_VFrameSet_HFrameSet_VFrameSet1_workFrame_04006015_form_div_work_div_search_rcd_calFrom_calendaredit_input"
            CAL_TO_ID   = "mainframe_VFrameSet_HFrameSet_VFrameSet1_workFrame_04006015_form_div_work_div_search_rcd_calTo_calendaredit_input"
            date_frame = None
            for frame in erp_page.frames:
                try:
                    if frame.locator(f"#{CAL_FROM_ID}").count() > 0:
                        date_frame = frame
                        print(f"      ✅ 날짜 입력 프레임 발견")
                        break
                except Exception:
                    continue
            if date_frame is None:
                date_frame = erp_page

            # 시작일 입력 (전체 선택 후 한번에 입력)
            date_from_str = DATE_FROM.replace("-", "")  # 20260201
            cal_from = date_frame.locator(f"#{CAL_FROM_ID}")
            cal_from.click()
            erp_page.wait_for_timeout(300)
            cal_from.press("Control+a")
            erp_page.wait_for_timeout(200)
            cal_from.press_sequentially(date_from_str, delay=150)
            erp_page.wait_for_timeout(300)
            cal_from.press("Tab")
            erp_page.wait_for_timeout(500)
            print(f"      ✅ 시작일: {DATE_FROM}")

            # 종료일 입력 (전체 선택 후 한번에 입력)
            date_to_str = DATE_TO.replace("-", "")  # 20260131
            cal_to = date_frame.locator(f"#{CAL_TO_ID}")
            cal_to.click()
            erp_page.wait_for_timeout(300)
            cal_to.press("Control+a")
            erp_page.wait_for_timeout(200)
            cal_to.press_sequentially(date_to_str, delay=150)
            erp_page.wait_for_timeout(300)
            cal_to.press("Tab")
            erp_page.wait_for_timeout(500)
            print(f"      ✅ 종료일: {DATE_TO}")

            # 판매구분 선택
            erp_page.locator(f"div:nth-child(3) > div > {WF}_div_search_rdo_sellGubun_item_radioimg > div").click()
            erp_page.wait_for_timeout(300)
            erp_page.locator("#mainframe").press("Enter")
            erp_page.wait_for_timeout(300)
            print(f"      ✅ 판매구분 선택")

            # 좌측 패널 숨기기
            erp_page.locator("#mainframe_VFrameSet_HFrameSet_leftShortFrame_form_btnShow > div").click()
            erp_page.wait_for_timeout(500)

            # 조회
            erp_page.get_by_text("조회", exact=True).click()
            print(f"      ⏳ 조회 결과 로딩 대기 중...")
            erp_page.get_by_text("1", exact=True).wait_for(timeout=30000)
            erp_page.wait_for_timeout(500)
            print(f"      ✅ 조회 완료")

            # ── STEP 5: 엑셀 다운로드 ─────────────────────────
            print(f"\n[5/5] 📥 엑셀 다운로드")
            erp_page.get_by_text("1", exact=True).click(button="right")
            erp_page.wait_for_timeout(2000)

            # 자료변환 클릭 후 로딩 대기 → 다운로드 시작
            with erp_page.expect_download(timeout=600000) as download_info:
                erp_page.locator(
                    "#mainframe_VFrameSet_HFrameSet_VFrameSet1_workFrame_04006015_form_div_work_div_workgrd_listPopupMenu_31TextBoxElement"
                ).get_by_text("자료변환 [Export File]").click()
                print(f"      ⏳ 다운로드 대기 중... (로딩 완료 후 자동 저장)")

            download = download_info.value
            save_path = os.path.join(DOWNLOAD_DIR, download.suggested_filename)
            download.save_as(save_path)
            print(f"      ✅ 다운로드 완료: {save_path}")
            return True

        except Exception as e:
            print(f"\n❌ ERP 오류 발생: {e}")
            traceback.print_exc()
            return False

        finally:
            browser.close()
            print("🏁 브라우저 종료")


# ============================================================
# 🚀 메인 실행
# ============================================================
def main():
    print("="*50)
    print("🤖 RPA 통합 자동화 v1.0")
    print("   PHASE 1: ERP 다운로드")
    print("   PHASE 2: 엑셀 가공")
    print(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    print("="*50)

    # PHASE 1: ERP 다운로드
    erp_ok = run_erp()

    if not erp_ok:
        print("\n❌ ERP 다운로드 실패 - 엑셀 가공을 중단합니다.")
        input("[Enter] to exit...")
        return

    # PHASE 2: 엑셀 가공
    excel_ok = run_excel()

    print("\n" + "="*50)
    if excel_ok:
        print("✅ 전체 RPA 완료!")
        print(f"   📁 OUTPUT: {OUTPUT_DIR}")
    else:
        print("⚠️  엑셀 가공 중 오류가 발생했습니다.")
    print("="*50)
    input("[Enter] to exit...")

if __name__ == "__main__":
    main()